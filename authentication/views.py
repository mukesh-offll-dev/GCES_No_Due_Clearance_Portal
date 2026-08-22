from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime, date, timedelta, timezone
from zoneinfo import ZoneInfo
import logging
from .decorators import institution_login_required
from .institution_users import INSTITUTION_USERS
from .mongo import institution_logs , students_col , no_due_col, portal_settings
from bson.errors import InvalidId
from bson import ObjectId
from pymongo.errors import DuplicateKeyError
from django.conf import settings
from .utils import save_receipt , reset_expired_no_dues
from .events import notify_student, notify_all_students, notify_office, notify_faculty, notify_all_offices
import re
import cloudinary.uploader
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook ,load_workbook

logger = logging.getLogger("nodue")
audit_logger = logging.getLogger("nodue.audit")

# Office roles that may approve/reject their own no-due requests.
_OFFICE_ROLES = ("LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT")

# ReportLab Imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas 


# =================== HELPER: ROLE → REDIRECT NAME ====================
_ROLE_REDIRECT = {
    "LIBRARY":    "library_dashboard",
    "HOSTEL":     "hostel_dashboard",
    "COLLEGE":    "college_dashboard",
    "FACULTY":    "faculty_dashboard",
    "DEPARTMENT": "department_dashboard",
    "STUDENT":    "student_dashboard",
}


def _safe_referer(request, role):
    """
    Redirect back to the referring page, but only if it's same-host (prevents
    open-redirect via a forged Referer). Falls back to the role dashboard.
    """
    from urllib.parse import urlparse
    referer = request.META.get("HTTP_REFERER")
    if referer:
        ref_host = urlparse(referer).netloc
        if not ref_host or ref_host == request.get_host():
            return redirect(referer)
    return redirect(_ROLE_REDIRECT.get(role, "index"))


# Reusable $lookup stage joining a no-due request to its student.
_STUDENT_LOOKUP = {
    "$lookup": {
        "from": "students",
        "localField": "student_id",
        "foreignField": "_id",
        "as": "student",
    }
}


def _office_branch_summary(office, branches):
    """
    Count PENDING requests per branch for one office in a SINGLE aggregation
    (was one aggregation per branch). Returns {branch: count} with 0 defaults.
    """
    pipeline = [
        {"$match": {"office": office, "status": "PENDING"}},
        _STUDENT_LOOKUP,
        {"$unwind": "$student"},
        {"$group": {"_id": "$student.branch", "count": {"$sum": 1}}},
    ]
    counts = {row["_id"]: row["count"] for row in no_due_col.aggregate(pipeline)}
    return {b: counts.get(b, 0) for b in branches}


def _office_year_summary(office, branch, years=(1, 2, 3, 4)):
    """
    Count PENDING requests per year for one office+branch in a SINGLE
    aggregation (was one aggregation per year). Returns {year: count}.
    """
    pipeline = [
        {"$match": {"office": office, "status": "PENDING"}},
        _STUDENT_LOOKUP,
        {"$unwind": "$student"},
        {"$match": {"student.branch": branch}},
        {"$group": {"_id": "$student.year", "count": {"$sum": 1}}},
    ]
    counts = {row["_id"]: row["count"] for row in no_due_col.aggregate(pipeline)}
    return {y: counts.get(y, 0) for y in years}


def _office_pending_requests(office, branch, year):
    """PENDING requests for an office filtered to branch+year, with a string id."""
    try:
        year_int = int(year)
    except (ValueError, TypeError):
        return []
    requests = list(no_due_col.aggregate([
        {"$match": {"office": office, "status": "PENDING"}},
        _STUDENT_LOOKUP,
        {"$unwind": "$student"},
        {"$match": {"student.branch": branch, "student.year": year_int}},
    ]))
    for r in requests:
        r["id"] = str(r["_id"])
    return requests


# ================= INDEX = INSTITUTION LOGIN =================
# ================= INDEX = INSTITUTION LOGIN =================
def index(request):
    # ── If a valid session already exists, skip the login page ──
    role = request.session.get("role")
    if role:
        dest = _ROLE_REDIRECT.get(role)
        if dest:
            return redirect(dest)

    student_error = request.session.pop("student_error", None)

    if request.method == "POST":
        office = request.POST.get("office", "").strip()
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        department = request.POST.get("department", "").strip()

        # ================= DEPARTMENT LOGIN =================
        if office == "department":
            dept = INSTITUTION_USERS.get("department", {}).get(department)

            if dept and dept.get("username") == username and dept.get("password") == password:
                request.session["role"] = "DEPARTMENT"
                request.session["department"] = department

                try:
                    institution_logs.insert_one({
                        "office": "DEPARTMENT",
                        "department": department,
                        "username": username,
                        "login_time": datetime.now()
                    })
                except Exception as exc:
                    logger.warning("Failed to record department login log: %s", exc)

                return redirect("department_dashboard")

        # ================= OTHER OFFICES =================
        else:
            office_data = INSTITUTION_USERS.get(office)

            if office_data and office_data.get("username") == username and office_data.get("password") == password:
                role = office_data["role"]
                request.session["role"] = role

                try:
                    institution_logs.insert_one({
                        "office": role,
                        "username": username,
                        "login_time": datetime.now()
                    })
                except Exception as exc:
                    logger.warning("Failed to record office login log: %s", exc)

                # 🔀 REDIRECT BASED ON ROLE
                dest = _ROLE_REDIRECT.get(role, "index")
                return redirect(dest)

        # ❌ INVALID LOGIN
        audit_logger.warning("LOGIN failed office=%s username=%s ip=%s",
                             office, username, request.META.get("REMOTE_ADDR"))
        messages.error(request, "Invalid credentials")

    return render(request, "index.html", {"student_error": student_error})


def student_login(request):
    # ── Already logged in as STUDENT? Skip login ──
    if request.session.get("role") == "STUDENT" and request.session.get("student_id"):
        return redirect("student_dashboard")

    if request.method == "POST":
        reg_no = request.POST.get("reg_no", "").strip()
        dob = request.POST.get("dob", "").strip()

        # ✅ VALIDATIONS
        if not reg_no.isdigit() or len(reg_no) != 12:
            request.session["student_error"] = "Register Number must be 12 digits"
            return redirect("index")

        try:
            datetime.strptime(dob, "%Y-%m-%d")
        except ValueError:
            request.session["student_error"] = "Invalid Date of Birth"
            return redirect("index")

        # 🔎 CHECK STUDENT IN DB
        try:
            student = students_col.find_one({
                "reg_no": reg_no,
                "dob": dob
            })
        except Exception as exc:
            logger.error("Database query failed on student_login: %s", exc)
            request.session["student_error"] = "Service temporarily unavailable. Please try again."
            return redirect("index")

        if not student:
            audit_logger.warning("STUDENT LOGIN failed reg_no=%s ip=%s",
                                 reg_no, request.META.get("REMOTE_ADDR"))
            request.session["student_error"] = "Invalid credentials"
            return redirect("index")

        # ================= LOGIN SUCCESS =================
        # 🔒 Cycle session ID (prevent session-fixation attacks).
        try:
            request.session.cycle_key()
        except Exception:
            request.session.flush()

        request.session["role"]       = "STUDENT"
        request.session["student_id"] = str(student["_id"])

        return redirect("student_dashboard")

    return redirect("index")


 

def check_no_due_access_status():
    settings_doc = portal_settings.find_one({"_id": "global_config"})
    if not settings_doc:
        return False
    enabled = settings_doc.get("no_due_access_enabled", False)
    if enabled:
        auto_disable_at = settings_doc.get("auto_disable_at")
        disabled_now = False
        if auto_disable_at:
            if auto_disable_at.tzinfo is None:
                auto_disable_at = auto_disable_at.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) >= auto_disable_at:
                portal_settings.update_one(
                    {"_id": "global_config"},
                    {"$set": {"no_due_access_enabled": False}}
                )
                disabled_now = True
        else:
            portal_settings.update_one(
                {"_id": "global_config"},
                {"$set": {"no_due_access_enabled": False}}
            )
            disabled_now = True

        if disabled_now:
            try:
                notify_all_students("no_due_access_toggled", {
                    "enabled": False,
                    "auto_disable_str": "",
                })
                notify_faculty("no_due_access_toggled", {
                    "enabled": False,
                    "auto_disable_str": "",
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast auto-disable: %s", be)
            return False
    return enabled


@institution_login_required
def student_dashboard(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    reset_expired_no_dues(no_due_col)

    student_id_raw = request.session.get("student_id")
    if not student_id_raw:
        request.session.flush()
        return redirect("index")

    try:
        student_id = ObjectId(student_id_raw)
        student = students_col.find_one({"_id": student_id})
    except Exception as exc:
        logger.error("Error fetching student %s: %s", student_id_raw, exc)
        request.session.flush()
        return redirect("index")

    if not student:
        request.session.flush()
        return redirect("index")

    # ── Determine offices based on student type ──
    student_type = student.get("student_type", "Hosteller")
    if student_type == "Day Scholar":
        offices = ["LIBRARY", "COLLEGE", "DEPARTMENT"]
        required_count = 3
    else:
        offices = ["LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT"]
        required_count = 4

    try:
        existing = {
            d.get("office"): d
            for d in no_due_col.find({"student_id": student_id})
            if d.get("office")
        }
    except Exception as exc:
        logger.error("Error querying dues for student %s: %s", student_id, exc)
        existing = {}

    dues = []
    all_approved = True   # 🔥 FLAG

    for office in offices:
        d = existing.get(office, {
            "office": office,
            "status": "NOT_SENT",
            "attempts_used": 0
        })

        attempts_used = d.get("attempts_used", 0)
        cooldown_expiry = d.get("cooldown_expiry")
        is_cooldown_active = False
        cooldown_expiry_iso = ""
        cooldown_remaining_seconds = 0

        if cooldown_expiry:
            if cooldown_expiry.tzinfo is None:
                cooldown_expiry = cooldown_expiry.replace(tzinfo=timezone.utc)
            now_utc = datetime.now(timezone.utc)
            if now_utc < cooldown_expiry:
                is_cooldown_active = True
                cooldown_expiry_iso = cooldown_expiry.isoformat()
                cooldown_remaining_seconds = int((cooldown_expiry - now_utc).total_seconds())

        d["attempts_used"] = attempts_used
        d["attempts_remaining"] = max(0, 2 - attempts_used)
        d["is_cooldown_active"] = is_cooldown_active
        d["cooldown_expiry_iso"] = cooldown_expiry_iso
        d["cooldown_remaining_seconds"] = cooldown_remaining_seconds

        dues.append(d)

        # 🔴 if ANY required office not approved → false
        if d.get("status") != "APPROVED":
            all_approved = False

    # Query promotion logs for this student
    logs = []
    try:
        from .mongo import promotion_logs
        logs = list(promotion_logs.find({"student_id": student_id}).sort("promotion_time", -1))
    except Exception as exc:
        logger.warning("Error loading promotion logs for student %s: %s", student_id, exc)

    no_due_access_enabled = check_no_due_access_status()

    return render(request, "student_dashboard.html", {
        "student": student,
        "dues": dues,
        "all_approved": all_approved,
        "promotion_logs": logs,
        "student_type": student_type,
        "no_due_access_enabled": no_due_access_enabled,
    })


@institution_login_required
def update_student_profile(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    if request.method == "POST":
        year = request.POST.get("year", "")
        semester = request.POST.get("semester", "")

        # 🔐 STRICT VALIDATION
        if not year.isdigit() or not (1 <= int(year) <= 4):
            return redirect("student_dashboard")

        if not semester.isdigit() or not (1 <= int(semester) <= 8):
            return redirect("student_dashboard")

        student_id_raw = request.session.get("student_id")
        if not student_id_raw:
            request.session.flush()
            return redirect("index")

        try:
            students_col.update_one(
                {"_id": ObjectId(student_id_raw)},
                {"$set": {
                    "year": int(year),
                    "semester": int(semester)
                }}
            )
        except Exception as exc:
            logger.error("Failed to update student profile: %s", exc)

    return redirect("student_dashboard")


@institution_login_required
def no_due_certificate(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    student_id_raw = request.session.get("student_id")
    if not student_id_raw:
        request.session.flush()
        return redirect("index")

    try:
        student_id = ObjectId(student_id_raw)
        student = students_col.find_one({"_id": student_id})
    except Exception:
        request.session.flush()
        return redirect("index")

    if not student:
        request.session.flush()
        return redirect("index")

    # Fetch all approved no-dues
    try:
        dues = list(no_due_col.find({
            "student_id": student_id,
            "status": "APPROVED"
        }))
    except Exception as exc:
        logger.error("Failed to fetch approved dues for certificate: %s", exc)
        dues = []

    # 🔐 Safety check
    student_type = student.get("student_type", "Hosteller")
    required_count = 3 if student_type == "Day Scholar" else 4
    if len(dues) < required_count:
        return redirect("student_dashboard")

    # Convert to simple dict for template
    no_dues_status = {
        "LIBRARY": "Completed",
        "COLLEGE": "Completed",
        "DEPARTMENT": "Completed"
    }
    if student_type == "Hosteller":
        no_dues_status["HOSTEL"] = "Completed"

    return render(request, "no_due_certificate.html", {
        "student": student,
        "no_dues": no_dues_status
    })


@institution_login_required
def send_hostel_request(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    if not check_no_due_access_status():
        err = "No Due process is currently locked. Please contact your Faculty."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
            return JsonResponse({"success": False, "error": err}, status=400)
        messages.error(request, err)
        return redirect("student_dashboard")

    if request.method == "POST":
        student_id_raw = request.session.get("student_id")
        if not student_id_raw:
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Unauthenticated"}, status=401)
            return redirect("index")

        try:
            student_id = ObjectId(student_id_raw)
        except (InvalidId, TypeError):
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Invalid session"}, status=401)
            return redirect("index")

        reset_expired_no_dues(no_due_col)

        existing_req = no_due_col.find_one({
            "student_id": student_id,
            "office": "HOSTEL"
        })

        if existing_req:
            cooldown_expiry = existing_req.get("cooldown_expiry")
            if cooldown_expiry:
                if cooldown_expiry.tzinfo is None:
                    cooldown_expiry = cooldown_expiry.replace(tzinfo=timezone.utc)
                if datetime.now(timezone.utc) < cooldown_expiry and existing_req.get("attempts_used", 0) >= 2:
                    err = "Request limit reached for Hostel Office. Cooldown active."
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                        return JsonResponse({"success": False, "error": err}, status=400)
                    messages.error(request, err)
                    return redirect("student_dashboard")

            if existing_req.get("attempts_used", 0) >= 2 and existing_req.get("status") == "REJECTED":
                err = "Maximum request limit (2 attempts) reached for Hostel Office."
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                    return JsonResponse({"success": False, "error": err}, status=400)
                messages.error(request, err)
                return redirect("student_dashboard")

        student = students_col.find_one({"_id": student_id})
        is_75 = student.get("is_75_scheme", False) if student else False

        receipt_url = None
        cloudinary_public_id = None

        if not is_75:
            if "receipt" in request.FILES:
                try:
                    upload = cloudinary.uploader.upload(
                        request.FILES["receipt"],
                        folder="no_dues/hostel",
                        resource_type="auto"
                    )
                    receipt_url = upload["secure_url"]
                    cloudinary_public_id = upload["public_id"]
                except Exception as exc:
                    logger.error("Cloudinary upload failed for student %s: %s",
                                 student_id, exc)
                    err = "Receipt upload failed. Please try again."
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                        return JsonResponse({"success": False, "error": err}, status=400)
                    messages.error(request, err)
                    return redirect("student_dashboard")
        else:
            receipt_url = None
            cloudinary_public_id = None

        current_attempts = existing_req.get("attempts_used", 0) if existing_req else 0
        new_attempts = min(2, current_attempts + 1)

        set_data = {
            "student_id": student_id,
            "office": "HOSTEL",
            "last_payment_id": "7.5 Scheme" if is_75 else request.POST.get("payment_id"),
            "status": "PENDING",
            "attempts_used": new_attempts,
            "created_at": datetime.now()
        }
        if receipt_url:
            set_data["receipt_url"] = receipt_url
        if cloudinary_public_id:
            set_data["cloudinary_public_id"] = cloudinary_public_id

        hostel_filter = {"student_id": student_id, "office": "HOSTEL"}
        try:
            no_due_col.update_one(hostel_filter, {"$set": set_data}, upsert=True)
        except DuplicateKeyError:
            # Concurrent submit created the doc first — apply our update to it.
            no_due_col.update_one(hostel_filter, {"$set": set_data})
        audit_logger.info("NO_DUE_REQUEST office=HOSTEL student=%s attempt=%d",
                          student_id, new_attempts)

        # 🚀 Broadcast real-time update to student and hostel dashboard
        try:
            notify_student(student_id, "request_status_updated", {
                "office": "HOSTEL",
                "status": "PENDING",
                "attempts_used": new_attempts,
                "attempts_remaining": max(0, 2 - new_attempts),
                "last_payment_id": set_data.get("last_payment_id"),
                "receipt_url": set_data.get("receipt_url"),
            })
            notify_office("HOSTEL", "new_request_submitted", {
                "student_id": str(student_id),
                "reg_no": student.get("reg_no") if student else "",
                "name": student.get("name") if student else "",
                "branch": student.get("branch") if student else "",
                "year": student.get("year") if student else 1,
                "semester": student.get("semester") if student else 1,
                "is_75_scheme": is_75,
                "last_payment_id": set_data.get("last_payment_id"),
                "receipt_url": set_data.get("receipt_url"),
                "attempts_used": new_attempts,
                "created_at": datetime.now().strftime("%d-%m-%Y %I:%M %p"),
            })
        except Exception as be:
            logger.warning("[WS] Failed to broadcast hostel request: %s", be)

        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
            return JsonResponse({"success": True, "status": "PENDING", "attempts_used": new_attempts})

    return redirect("student_dashboard")



@institution_login_required
def hostel_dashboard(request):
    if request.session.get("role") != "HOSTEL":
        return redirect("index")

    reset_expired_no_dues(no_due_col)

    branch = request.GET.get("branch")
    year = request.GET.get("year")
    semester = request.GET.get("semester")

    requests = []
    count = 0
    year_summary = {}
    branch_summary = {}

    if branch and year:
        requests = _office_pending_requests("HOSTEL", branch, year)
        count = len(requests)
    elif branch:
        year_summary = _office_year_summary("HOSTEL", branch)
    else:
        branch_summary = _office_branch_summary(
            "HOSTEL", ["CSE", "ECE", "EEE", "CIVIL", "MECH", "MCT"])

    return render(request, "hostel_dashboard.html", {
        "requests": requests,
        "count": count,
        "branch": branch,
        "year": year,
        "semester": semester,
        "branches": ["CSE","ECE","EEE","CIVIL","MECH","MCT"],
        "branch_summary": branch_summary,
        "year_summary": year_summary
    })


 

# ================= LIBRARY PAGE =================
@institution_login_required
def library_dashboard(request):
    if request.session.get("role") != "LIBRARY":
        return redirect("index")

    reset_expired_no_dues(no_due_col)

    branch = request.GET.get("branch")
    year = request.GET.get("year")
    semester = request.GET.get("semester")

    requests = []
    count = 0
    branch_summary = {}
    year_summary = {}

    if branch and year:
        requests = _office_pending_requests("LIBRARY", branch, year)
        count = len(requests)
    elif branch:
        year_summary = _office_year_summary("LIBRARY", branch)
    else:
        branch_summary = _office_branch_summary(
            "LIBRARY", ["CSE", "ECE", "EEE", "CIVIL", "MECH", "MCT"])

    return render(request, "institution_dashboard.html", {
        "office_name": "Library Office",
        "requests": requests,
        "branch": branch,
        "year": year,
        "semester": semester,
        "count": count,
        "branches": ["CSE","ECE","EEE","CIVIL","MECH","MCT"],
        "branch_summary": branch_summary,
        "year_summary": year_summary
    })




@institution_login_required
def bulk_approve(request):
    role = request.session.get("role")
    # 🔐 Authorization: only an office may approve, and only its OWN requests.
    if role not in _OFFICE_ROLES:
        return redirect("index")

    if request.method != "POST":
        return redirect(_ROLE_REDIRECT.get(role, "index"))

    ids = request.POST.getlist("request_ids")
    object_ids = []
    for i in ids:
        try:
            object_ids.append(ObjectId(i))
        except (InvalidId, TypeError):
            pass  # ignore malformed ids instead of 500-ing

    approved = 0
    if object_ids:
        # Pre-fetch requests to know which students are being approved
        pending_to_approve = list(no_due_col.find(
            {"_id": {"$in": object_ids}, "office": role, "status": "PENDING"},
            {"student_id": 1, "office": 1}
        ))

        # Scope by office==role AND status==PENDING so:
        #   • one office cannot approve another office's request,
        #   • already-approved/rejected requests are never re-processed
        #     (idempotent — protects against duplicate submits / double clicks).
        result = no_due_col.update_many(
            {"_id": {"$in": object_ids}, "office": role, "status": "PENDING"},
            {"$set": {"status": "APPROVED", "updated_at": datetime.now()}}
        )
        approved = result.modified_count
        audit_logger.info("APPROVE role=%s requested=%d approved=%d",
                          role, len(object_ids), approved)

        # 🚀 Broadcast real-time approval to each student individually
        student_statuses = {}
        for r in pending_to_approve:
            sid = r.get("student_id")
            sid_str = str(sid)
            student_statuses[sid_str] = "Completed"
            try:
                notify_student(sid, "request_status_updated", {
                    "office": role,
                    "status": "APPROVED",
                    "updated_at": datetime.now().isoformat(),
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast APPROVED to student %s: %s", sid_str, be)
        try:
            notify_office(role, "requests_processed", {
                "processed_ids": [str(i) for i in object_ids],
                "student_ids": list(student_statuses.keys()),
                "student_statuses": student_statuses,
                "action": "APPROVED",
                "count": approved,
            })
        except Exception as be:
            logger.warning("[WS] Failed to broadcast approval to office %s: %s", role, be)

    return _safe_referer(request, role)



@institution_login_required
def reject_request(request):
    role = request.session.get("role")
    # 🔐 Authorization: only an office may reject, and only its OWN requests.
    if role not in _OFFICE_ROLES:
        return redirect("index")

    if request.method != "POST":
        return redirect(_ROLE_REDIRECT.get(role, "index"))

    req_ids = request.POST.getlist("request_ids")
    if not req_ids:
        req_id = request.POST.get("req_id")
        if req_id:
            req_ids = [req_id]
        else:
            req_ids = []

    reason = request.POST.get("reason")

    if req_ids:
        object_ids = []
        for rid in req_ids:
            try:
                object_ids.append(ObjectId(rid))
            except (InvalidId, TypeError):
                pass

        # 🔥 If hostel + file exists → delete from Cloudinary
        # Scope by office==role AND status==PENDING (idempotent, cross-office safe).
        requests_to_reject = list(no_due_col.find(
            {"_id": {"$in": object_ids}, "office": role, "status": "PENDING"}
        )) if object_ids else []
        now_utc = datetime.now(timezone.utc)
        now_naive = datetime.now()

        student_statuses = {}
        for req in requests_to_reject:
            if req.get("office") == "HOSTEL":
                public_id = req.get("cloudinary_public_id")
                if public_id:
                    try:
                        cloudinary.uploader.destroy(
                            public_id,
                            resource_type="raw"
                        )
                    except Exception:
                        pass
                    try:
                        cloudinary.uploader.destroy(
                            public_id,
                            resource_type="image"
                        )
                    except Exception:
                        pass

            attempts_used = req.get("attempts_used", 1)
            update_payload = {
                "status": "REJECTED",
                "reject_reason": reason,
                "receipt_url": None,
                "cloudinary_public_id": None,
                "updated_at": now_naive
            }

            if attempts_used >= 2:
                update_payload["second_rejection_at"] = now_utc
                update_payload["cooldown_expiry"] = now_utc + timedelta(hours=24)

            no_due_col.update_one(
                {"_id": req["_id"], "status": "PENDING"},
                {"$set": update_payload}
            )

            sid_str = str(req.get("student_id"))
            student_statuses[sid_str] = "Incomplete"

            # 🚀 Broadcast real-time rejection to student
            try:
                notify_student(req.get("student_id"), "request_status_updated", {
                    "office": role,
                    "status": "REJECTED",
                    "reject_reason": reason,
                    "attempts_used": attempts_used,
                    "attempts_remaining": max(0, 2 - attempts_used),
                    "is_cooldown_active": (attempts_used >= 2),
                    "cooldown_expiry_iso": update_payload.get("cooldown_expiry", "").isoformat() if update_payload.get("cooldown_expiry") else "",
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast REJECTED to student %s: %s", req.get("student_id"), be)

        audit_logger.info("REJECT role=%s count=%d", role, len(requests_to_reject))

        # 🚀 Broadcast processed removal to office dashboard
        try:
            notify_office(role, "requests_processed", {
                "processed_ids": [str(r["_id"]) for r in requests_to_reject],
                "student_ids": [str(r.get("student_id")) for r in requests_to_reject],
                "student_statuses": student_statuses,
                "action": "REJECTED",
                "count": len(requests_to_reject),
            })
        except Exception as be:
            logger.warning("[WS] Failed to broadcast rejection to office %s: %s", role, be)

    return _safe_referer(request, role)


@institution_login_required
def college_dashboard(request):
    if request.session.get("role") != "COLLEGE":
        return redirect("index")

    reset_expired_no_dues(no_due_col)

    branch = request.GET.get("branch")
    year = request.GET.get("year")
    semester = request.GET.get("semester")

    requests = []
    count = 0
    branch_summary = {}
    year_summary = {}

    if branch and year:
        requests = _office_pending_requests("COLLEGE", branch, year)
        count = len(requests)
    elif branch:
        year_summary = _office_year_summary("COLLEGE", branch)
    else:
        branch_summary = _office_branch_summary(
            "COLLEGE", ["CSE", "ECE", "EEE", "CIVIL", "MECH", "MCT"])

    return render(request, "institution_dashboard.html", {
        "office_name": "College Office",
        "requests": requests,
        "branch": branch,
        "year": year,
        "semester": semester,
        "count": count,
        "branches": ["CSE","ECE","EEE","CIVIL","MECH","MCT"],
        "branch_summary": branch_summary,
        "year_summary": year_summary
    })


@institution_login_required
def department_dashboard(request):
    if request.session.get("role") != "DEPARTMENT":
        return redirect("index")

    reset_expired_no_dues(no_due_col)

    dept = request.session.get("department")
    if not dept:
        return redirect("index")

    year = request.GET.get("year", "").strip()
    semester = request.GET.get("semester", "").strip()

    requests = []
    count = 0
    year_summary = {}

    if year:
        requests = _office_pending_requests("DEPARTMENT", dept, year)
        count = len(requests)
    else:
        year_summary = _office_year_summary("DEPARTMENT", dept)

    return render(request, "institution_dashboard.html", {
        "office_name": "Department",
        "requests": requests,
        "branch": dept,
        "year": year,
        "semester": semester,
        "count": count,
        "year_summary": year_summary
    })


@institution_login_required
def send_no_due_request(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    if not check_no_due_access_status():
        err = "No Due process is currently locked. Please contact your Faculty."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
            return JsonResponse({"success": False, "error": err}, status=400)
        messages.error(request, err)
        return redirect("student_dashboard")

    if request.method == "POST":
        office = request.POST.get("office", "").strip()
        student_id_raw = request.session.get("student_id")
        if not student_id_raw:
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Unauthenticated"}, status=401)
            return redirect("index")

        try:
            student_id = ObjectId(student_id_raw)
        except (InvalidId, TypeError):
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Invalid session"}, status=401)
            return redirect("index")

        reset_expired_no_dues(no_due_col)

        existing_req = no_due_col.find_one({
            "student_id": student_id,
            "office": office
        })

        if existing_req:
            cooldown_expiry = existing_req.get("cooldown_expiry")
            if cooldown_expiry:
                if cooldown_expiry.tzinfo is None:
                    cooldown_expiry = cooldown_expiry.replace(tzinfo=timezone.utc)
                if datetime.now(timezone.utc) < cooldown_expiry and existing_req.get("attempts_used", 0) >= 2:
                    err = f"Request limit reached for {office.title() if office else ''} Office. Cooldown active."
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                        return JsonResponse({"success": False, "error": err}, status=400)
                    messages.error(request, err)
                    return redirect("student_dashboard")

            if existing_req.get("attempts_used", 0) >= 2 and existing_req.get("status") == "REJECTED":
                err = f"Maximum request limit (2 attempts) reached for {office.title() if office else ''} Office."
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                    return JsonResponse({"success": False, "error": err}, status=400)
                messages.error(request, err)
                return redirect("student_dashboard")

        current_attempts = existing_req.get("attempts_used", 0) if existing_req else 0
        new_attempts = min(2, current_attempts + 1)

        data = {
            "student_id": student_id,
            "office": office,
            "status": "PENDING",
            "attempts_used": new_attempts,
            "created_at": datetime.now()
        }

        # HOSTEL extra fields
        if office == "HOSTEL":
            if "receipt" in request.FILES:
                data["receipt"] = save_receipt(request.FILES["receipt"])
            data["last_payment_id"] = request.POST.get("payment_id")

        # 🔁 UPDATE if exists, else INSERT
        office_filter = {"student_id": data["student_id"], "office": office}
        try:
            no_due_col.update_one(office_filter, {"$set": data}, upsert=True)
        except DuplicateKeyError:
            # Concurrent submit created the doc first — apply our update to it.
            no_due_col.update_one(office_filter, {"$set": data})
        audit_logger.info("NO_DUE_REQUEST office=%s student=%s attempt=%d",
                          office, student_id, new_attempts)

        # 🚀 Broadcast real-time event to student and office dashboard
        try:
            student = students_col.find_one({"_id": student_id})
            notify_student(student_id, "request_status_updated", {
                "office": office,
                "status": "PENDING",
                "attempts_used": new_attempts,
                "attempts_remaining": max(0, 2 - new_attempts),
            })
            notify_office(office, "new_request_submitted", {
                "student_id": str(student_id),
                "reg_no": student.get("reg_no") if student else "",
                "name": student.get("name") if student else "",
                "branch": student.get("branch") if student else "",
                "year": student.get("year") if student else 1,
                "semester": student.get("semester") if student else 1,
                "is_75_scheme": student.get("is_75_scheme", False) if student else False,
                "attempts_used": new_attempts,
                "created_at": datetime.now().strftime("%d-%m-%Y %I:%M %p"),
            }, department=student.get("branch") if office == "DEPARTMENT" and student else None)
        except Exception as be:
            logger.warning("[WS] Failed to broadcast send_no_due_request: %s", be)

        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
            return JsonResponse({"success": True, "status": "PENDING", "attempts_used": new_attempts})

    return redirect("student_dashboard")


@institution_login_required
def retry_request(request):
    if request.session.get("role") != "STUDENT":
        return redirect("index")

    if not check_no_due_access_status():
        err = "No Due process is currently locked. Please contact your Faculty."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
            return JsonResponse({"success": False, "error": err}, status=400)
        messages.error(request, err)
        return redirect("student_dashboard")

    if request.method == "POST":
        office = request.POST.get("office", "").strip()
        student_id_raw = request.session.get("student_id")
        if not student_id_raw:
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Unauthenticated"}, status=401)
            return redirect("index")

        try:
            student_id = ObjectId(student_id_raw)
        except (InvalidId, TypeError):
            request.session.flush()
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Invalid session"}, status=401)
            return redirect("index")

        reset_expired_no_dues(no_due_col)

        existing_req = no_due_col.find_one({
            "student_id": student_id,
            "office": office,
            "status": "REJECTED"
        })

        if existing_req:
            cooldown_expiry = existing_req.get("cooldown_expiry")
            if cooldown_expiry:
                if cooldown_expiry.tzinfo is None:
                    cooldown_expiry = cooldown_expiry.replace(tzinfo=timezone.utc)
                if datetime.now(timezone.utc) < cooldown_expiry and existing_req.get("attempts_used", 0) >= 2:
                    err = f"Request limit reached for {office.title() if office else ''} Office. Cooldown active."
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                        return JsonResponse({"success": False, "error": err}, status=400)
                    messages.error(request, err)
                    return redirect("student_dashboard")

            if existing_req.get("attempts_used", 0) >= 2:
                err = f"Maximum request limit (2 attempts) reached for {office.title() if office else ''} Office."
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                    return JsonResponse({"success": False, "error": err}, status=400)
                messages.error(request, err)
                return redirect("student_dashboard")

            no_due_col.update_one(
                {
                    "_id": existing_req["_id"]
                },
                {
                    "$set": {
                        "status": "NOT_SENT",
                        "created_at": datetime.now(),
                        "updated_at": datetime.now()
                    },
                    "$unset": {
                        "reject_reason": ""
                    }
                }
            )

            # 🚀 Broadcast real-time update to student
            try:
                attempts_used = existing_req.get("attempts_used", 1)
                notify_student(student_id, "request_status_updated", {
                    "office": office,
                    "status": "NOT_SENT",
                    "attempts_used": attempts_used,
                    "attempts_remaining": max(0, 2 - attempts_used),
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast retry_request: %s", be)

            if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.POST.get("format") == "json":
                return JsonResponse({"success": True, "status": "NOT_SENT", "attempts_used": attempts_used})

    return redirect("student_dashboard")






@institution_login_required
def faculty_dashboard(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")
    reset_expired_no_dues(no_due_col)
    
    add_error = request.session.pop("add_error", None)
    add_success = request.session.pop("add_success", None)

    branch = request.GET.get("branch", "").strip()
    year = request.GET.get("year", "").strip()

    students = []
    count = 0

    if branch and year:
        try:
            year_int = int(year)
        except (ValueError, TypeError):
            year_int = None

        if year_int is not None:
            try:
                student_docs = list(students_col.find({
                    "branch": branch,
                    "year": year_int
                }))
            except Exception as exc:
                logger.error("Failed to query students in faculty_dashboard: %s", exc)
                student_docs = []
        else:
            student_docs = []

        # 🔥 Batch-fetch ALL no-dues for these students in ONE query (no N+1).
        student_ids = [s["_id"] for s in student_docs if "_id" in s]
        dues_by_student = {}
        if student_ids:
            try:
                for d in no_due_col.find({"student_id": {"$in": student_ids}}):
                    sid = d.get("student_id")
                    office = d.get("office")
                    status = d.get("status", "NOT_SENT")
                    if sid and office:
                        dues_by_student.setdefault(sid, {})[office] = status
            except Exception as exc:
                logger.error("Failed to query dues in faculty_dashboard: %s", exc)

        for s in student_docs:
            student_id = s.get("_id")
            if not student_id:
                continue

            no_dues = {
                "LIBRARY": "NOT_SENT",
                "HOSTEL": "NOT_SENT",
                "COLLEGE": "NOT_SENT",
                "DEPARTMENT": "NOT_SENT"
            }
            for office, status in dues_by_student.get(student_id, {}).items():
                no_dues[office] = status

            s_type = s.get("student_type", "Hosteller")

            students.append({
                "id": str(student_id),
                "roll_no": s.get("roll_no", ""),
                "reg_no": s.get("reg_no", ""),
                "name": s.get("name", ""),
                "semester": s.get("semester", ""),
                "dob": s.get("dob", ""),
                "phone": s.get("phone", ""),
                "branch": s.get("branch", ""),
                "year": s.get("year", ""),
                "student_type": s_type,
                "is_75_scheme": s.get("is_75_scheme", False),
                "no_dues": no_dues
            })

        count = len(students)

    return render(request, "faculty_dashboard.html", {
        "students": students,
        "count": count,
        "branch": branch,
        "year": year,
        "add_error": add_error,
        "add_success": add_success
    })


@institution_login_required
def update_75_scheme(request):
    if request.session.get("role") != "FACULTY":
        messages.error(request, "Unauthorized access.")
        return redirect("index")

    if request.method == "POST":
        branch = request.POST.get("branch", "").strip()
        year = request.POST.get("year", "").strip()

        selected_ids_str = request.POST.getlist("selected_student_ids")
        selected_ids = []
        for sid in selected_ids_str:
            try:
                selected_ids.append(ObjectId(sid))
            except Exception:
                pass

        query = {}
        if year and year.isdigit():
            query["year"] = int(year)
        if branch:
            query["branch"] = branch

        if query:
            try:
                # Set is_75_scheme = False for all students in this cohort
                students_col.update_many(query, {"$set": {"is_75_scheme": False}})
                # Set is_75_scheme = True for selected students
                if selected_ids:
                    students_col.update_many({"_id": {"$in": selected_ids}}, {"$set": {"is_75_scheme": True}})
                messages.success(request, "7.5 Scheme student status updated successfully.")

                # 🚀 Broadcast 7.5 Scheme updates
                try:
                    for sid in selected_ids:
                        notify_student(sid, "scheme_75_updated", {"is_75_scheme": True})
                    notify_faculty("scheme_75_updated", {"branch": branch, "year": year})
                    notify_office("HOSTEL", "scheme_75_updated", {"branch": branch, "year": year})
                except Exception as be:
                    logger.warning("[WS] Failed to broadcast 7.5 scheme update: %s", be)
            except Exception as exc:
                logger.error("Failed to update 7.5 scheme status: %s", exc)
                messages.error(request, "Failed to update 7.5 scheme status.")

        redirect_url = f"/faculty/?branch={branch or ''}&year={year or ''}"
        return redirect(redirect_url)

    return redirect("faculty_dashboard")


@institution_login_required
def add_student(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if request.method == "POST":
        roll_no = request.POST.get("roll_no", "").strip()
        reg_no = request.POST.get("reg_no", "").strip()
        name = request.POST.get("name", "").strip()
        dob = request.POST.get("dob", "").strip()
        semester = request.POST.get("semester", "").strip()
        phone = request.POST.get("phone", "").strip()
        branch = request.POST.get("branch", "").strip()
        year = request.POST.get("year", "").strip()

        # ================= FORMAT FIXES =================

        # 🔥 NAME → FULL CAPS
        name = name.upper()

        # 🔥 ROLL NO → department code CAPS (23cs533 → 23CS533)
        roll_no = re.sub(r'([a-zA-Z]+)', lambda m: m.group(1).upper(), roll_no)

        # ================= VALIDATIONS =================

        # Register Number → exactly 12 digits
        if not reg_no.isdigit() or len(reg_no) != 12:
            request.session["add_error"] = "Register Number must be exactly 12 digits"
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # Phone Number → exactly 10 digits
        if not phone.isdigit() or len(phone) != 10:
            request.session["add_error"] = "Phone Number must be exactly 10 digits"
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # DOB → valid date
        try:
            datetime.strptime(dob, "%Y-%m-%d")
        except ValueError:
            request.session["add_error"] = "Invalid Date of Birth format"
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # Semester → 1 to 8
        if not semester.isdigit() or not (1 <= int(semester) <= 8):
            request.session["add_error"] = "Semester must be between 1 and 8"
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # Year → 1 to 4
        if not year.isdigit() or not (1 <= int(year) <= 4):
            request.session["add_error"] = "Year must be between 1 and 4"
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # Duplicate check (Roll No / Reg No)
        try:
            existing_student = students_col.find_one({
                "$or": [
                    {"roll_no": roll_no},
                    {"reg_no": reg_no}
                ]
            })
        except Exception as exc:
            logger.error("Error checking duplicate student: %s", exc)
            existing_student = None

        if existing_student:
            request.session["add_error"] = (
                "Student with this Roll No or Register No already exists"
            )
            return redirect(f"/faculty/?branch={branch}&year={year}")

        # ================= INSERT =================
        student_type = request.POST.get("student_type", "Hosteller")
        if student_type not in ("Hosteller", "Day Scholar"):
            student_type = "Hosteller"

        try:
            inserted_student = students_col.insert_one({
                "roll_no": roll_no,
                "reg_no": reg_no,
                "name": name,
                "dob": dob,
                "phone": phone,
                "branch": branch,
                "year": int(year),
                "semester": int(semester),
                "student_type": student_type
            })
            request.session["add_success"] = "Student added successfully"
            try:
                notify_faculty("student_roster_updated", {"branch": branch, "year": year})
                notify_all_offices("student_roster_updated", {"branch": branch, "year": year})
            except Exception:
                pass
        except Exception as exc:
            logger.error("Failed to insert student: %s", exc)
            request.session["add_error"] = "Failed to add student. Please try again."

    return redirect(f"/faculty/?branch={branch}&year={year}")


@institution_login_required
def delete_students(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if request.method == "POST":
        ids = request.POST.getlist("student_ids")

        valid_object_ids = []
        for i in ids:
            try:
                valid_object_ids.append(ObjectId(i))
            except (InvalidId, TypeError):
                pass   # ignore empty / invalid ids

        if valid_object_ids:
            try:
                students_col.delete_many({
                    "_id": {"$in": valid_object_ids}
                })
                no_due_col.delete_many({
                    "student_id": {"$in": valid_object_ids}
                })
                try:
                    notify_faculty("student_roster_updated", {})
                    notify_all_offices("student_roster_updated", {})
                except Exception:
                    pass
            except Exception as exc:
                logger.error("Failed to delete students: %s", exc)

    branch = request.POST.get("branch", "").strip()
    year = request.POST.get("year", "").strip()

    return redirect(f"/faculty/?branch={branch}&year={year}")


@institution_login_required
def edit_student(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    branch = request.POST.get("branch", "").strip()
    year = request.POST.get("year", "").strip()

    if request.method == "POST":
        student_id = request.POST.get("student_id", "").strip()

        student_type = request.POST.get("student_type", "Hosteller")
        if student_type not in ("Hosteller", "Day Scholar"):
            student_type = "Hosteller"

        year_val_raw = request.POST.get("year", "").strip()
        sem_val_raw = request.POST.get("semester", "").strip()

        try:
            year_val = int(year_val_raw) if year_val_raw.isdigit() else 1
            sem_val = int(sem_val_raw) if sem_val_raw.isdigit() else 1

            students_col.update_one(
                {"_id": ObjectId(student_id)},
                {"$set": {
                    "roll_no": request.POST.get("roll_no", "").strip().upper(),
                    "reg_no": request.POST.get("reg_no", "").strip(),
                    "name": request.POST.get("name", "").strip().upper(),
                    "dob": request.POST.get("dob", "").strip(),
                    "year": year_val,
                    "phone": request.POST.get("phone", "").strip(),
                    "semester": sem_val,
                    "student_type": student_type
                }}
            )
            try:
                notify_student(student_id, "student_profile_updated", {
                    "roll_no": request.POST.get("roll_no", "").strip().upper(),
                    "name": request.POST.get("name", "").strip().upper(),
                    "year": year_val,
                    "semester": sem_val,
                    "student_type": student_type,
                })
                notify_faculty("student_roster_updated", {"branch": branch, "year": year})
            except Exception:
                pass
        except (InvalidId, Exception) as exc:
            logger.warning("Failed to edit student %s: %s", student_id, exc)

    return redirect(f"/faculty/?branch={branch}&year={year}")


@institution_login_required
def faculty_promotion_page(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    # Handle password verification
    if request.method == "POST":
        password = request.POST.get("promotion_password", "").strip()
        if password == "gces8301":
            request.session["promotion_unlocked"] = True
            return redirect("faculty_promotion")
        else:
            return render(request, "faculty_promotion_login.html", {"error": True})

    # Render login page if session is locked
    if not request.session.get("promotion_unlocked"):
        return render(request, "faculty_promotion_login.html")

    # Sync toggle expiration on load
    check_no_due_access_status()

    sem8_count = students_col.count_documents({"semester": 8})

    sem_counts = {}
    for sem in range(1, 9):
        sem_counts[sem] = students_col.count_documents({"semester": sem})

    global_sem8_count = sem8_count

    settings_doc = portal_settings.find_one({"_id": "global_config"})
    no_due_access_enabled = False
    enabled_at_str = ""
    auto_disable_at_str = ""
    auto_disable_at_iso = ""
    duration_days = 75
    remaining_seconds = 0

    if settings_doc:
        no_due_access_enabled = settings_doc.get("no_due_access_enabled", False)
        enabled_at = settings_doc.get("enabled_at")
        auto_disable_at = settings_doc.get("auto_disable_at")
        duration_days = settings_doc.get("duration_days", 75)
        
        tz_kolkata = ZoneInfo("Asia/Kolkata")
        
        if enabled_at:
            if enabled_at.tzinfo is None:
                enabled_at = enabled_at.replace(tzinfo=timezone.utc)
            enabled_at_kolkata = enabled_at.astimezone(tz_kolkata)
            enabled_at_str = enabled_at_kolkata.strftime("%d-%m-%Y %I:%M %p")
        if auto_disable_at:
            if auto_disable_at.tzinfo is None:
                auto_disable_at = auto_disable_at.replace(tzinfo=timezone.utc)
            auto_disable_at_kolkata = auto_disable_at.astimezone(tz_kolkata)
            auto_disable_at_str = auto_disable_at_kolkata.strftime("%d-%m-%Y %I:%M %p")
            auto_disable_at_iso = auto_disable_at_kolkata.isoformat()
            
            now_utc = datetime.now(timezone.utc)
            if auto_disable_at > now_utc:
                remaining_seconds = int((auto_disable_at - now_utc).total_seconds())

    return render(request, "faculty_promotion.html", {
        "sem8_count": sem8_count,
        "global_sem8_count": global_sem8_count,
        "sem_counts": sem_counts,
        "no_due_access_enabled": no_due_access_enabled,
        "enabled_at_str": enabled_at_str,
        "auto_disable_at_str": auto_disable_at_str,
        "auto_disable_at_iso": auto_disable_at_iso,
        "duration_days": duration_days,
        "remaining_seconds": remaining_seconds,
    })


@institution_login_required
def toggle_no_due_access(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if not request.session.get("promotion_unlocked"):
        messages.error(request, "Access denied. Please verify password first.")
        return redirect("faculty_promotion")

    if request.method == "POST":
        current_status = request.POST.get("current_status", "true") == "true"
        new_status = not current_status
        if new_status:
            duration_type = request.POST.get("duration_type", "").strip()
            custom_datetime_str = request.POST.get("custom_datetime", "").strip()

            tz_kolkata = ZoneInfo("Asia/Kolkata")
            now_utc = datetime.now(timezone.utc)
            now_kolkata = now_utc.astimezone(tz_kolkata)
            
            days = 0
            if duration_type == "recommended":
                auto_disable_at = now_utc + timedelta(days=75)
                days = 75
            elif duration_type == "custom":
                if not custom_datetime_str:
                    messages.error(request, "Auto Disable Date and Time is required to enable No Due Access.")
                    return redirect("faculty_promotion")
                try:
                    # Parse local datetime-local format: YYYY-MM-DDTHH:MM
                    naive_dt = datetime.fromisoformat(custom_datetime_str)
                    auto_disable_at_kolkata = naive_dt.replace(tzinfo=tz_kolkata)
                    auto_disable_at = auto_disable_at_kolkata.astimezone(timezone.utc)
                except ValueError:
                    messages.error(request, "Invalid Date and Time format.")
                    return redirect("faculty_promotion")

                if auto_disable_at <= now_utc:
                    messages.error(request, "Auto Disable Date and Time must be in the future.")
                    return redirect("faculty_promotion")

                delta = auto_disable_at - now_utc
                days = delta.days if delta.days > 0 else 1
            else:
                messages.error(request, "Auto Disable Duration is required to enable No Due Access.")
                return redirect("faculty_promotion")

            portal_settings.update_one(
                {"_id": "global_config"},
                {"$set": {
                    "no_due_access_enabled": True,
                    "enabled_at": now_utc,
                    "auto_disable_at": auto_disable_at,
                    "duration_days": days
                }},
                upsert=True
            )
            
            # Format display string for success message
            auto_disable_at_kolkata = auto_disable_at.astimezone(tz_kolkata)
            exp_str = auto_disable_at_kolkata.strftime("%d-%m-%Y %I:%M %p")
            messages.success(request, f"No Due Access has been successfully Enabled globally until {exp_str}.")

            # 🚀 Broadcast live enable event to all connected students and faculty
            try:
                notify_all_students("no_due_access_toggled", {
                    "enabled": True,
                    "auto_disable_str": exp_str,
                })
                notify_faculty("no_due_access_toggled", {
                    "enabled": True,
                    "auto_disable_str": exp_str,
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast no_due_access enable: %s", be)
        else:
            portal_settings.update_one(
                {"_id": "global_config"},
                {"$set": {"no_due_access_enabled": False}},
                upsert=True
            )
            messages.success(request, "No Due Access has been successfully Disabled globally.")

            # 🚀 Broadcast live disable event to all connected students and faculty
            try:
                notify_all_students("no_due_access_toggled", {
                    "enabled": False,
                    "auto_disable_str": "",
                })
                notify_faculty("no_due_access_toggled", {
                    "enabled": False,
                    "auto_disable_str": "",
                })
            except Exception as be:
                logger.warning("[WS] Failed to broadcast no_due_access disable: %s", be)

    return redirect("faculty_promotion")


@institution_login_required
def promote_students(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if not request.session.get("promotion_unlocked"):
        messages.error(request, "Access denied. Please verify password first.")
        return redirect("faculty_promotion")

    if request.method == "POST":
        # Ensure config document exists
        portal_settings.update_one(
            {"_id": "global_config"},
            {"$setOnInsert": {"promotion_in_progress": False}},
            upsert=True
        )

        # Acquire lock to prevent duplicate concurrent promotions
        lock_acquired = portal_settings.find_one_and_update(
            {"_id": "global_config", "promotion_in_progress": {"$ne": True}},
            {"$set": {"promotion_in_progress": True}}
        )
        if not lock_acquired:
            audit_logger.warning("PROMOTION blocked: another run in progress")
            messages.error(request, "A promotion operation is already in progress. Please wait.")
            return redirect("faculty_promotion")

        try:
            from_sem_raw = request.POST.get("from_semester")
            from_sem = int(from_sem_raw) if from_sem_raw and from_sem_raw.isdigit() else None

            if from_sem is not None:
                if not (1 <= from_sem <= 7):
                    messages.error(request, "Semester must be between 1 and 7.")
                    return redirect("faculty_promotion")
                students = list(students_col.find({"semester": from_sem}))
                if not students:
                    messages.warning(request, f"No students in Semester {from_sem} found for promotion.")
                    return redirect("faculty_promotion")
            else:
                students = list(students_col.find({"semester": {"$in": [1, 2, 3, 4, 5, 6, 7]}}))
                if not students:
                    messages.warning(request, "No students in Semesters 1 to 7 found for promotion.")
                    return redirect("faculty_promotion")

            progression = {
                1: (2, 0),
                2: (3, 1),
                3: (4, 0),
                4: (5, 1),
                5: (6, 0),
                6: (7, 1),
                7: (8, 0),
            }

            promoted_count = 0
            from .mongo import promotion_logs
            now = datetime.now()

            # Process updates student-by-student to maintain exact logs and individual status updates.
            for student in students:
                student_id = student["_id"]
                current_sem = student.get("semester")
                current_year = student.get("year", 1)

                if current_sem not in progression:
                    continue

                next_sem, year_change = progression[current_sem]
                new_year = current_year + year_change

                # Update student document
                students_col.update_one(
                    {"_id": student_id},
                    {"$set": {
                        "semester": next_sem,
                        "year": new_year
                    }}
                )

                student_type = student.get("student_type", "Hosteller")
                offices = ["LIBRARY", "COLLEGE", "DEPARTMENT"] if student_type == "Day Scholar" else ["LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT"]
                approved_count = no_due_col.count_documents({
                    "student_id": student_id,
                    "office": {"$in": offices},
                    "status": "APPROVED"
                })
                no_due_cleared = (approved_count == len(offices))

                # Reset dues
                no_due_col.update_many(
                    {"student_id": student_id},
                    {
                        "$set": {
                            "status": "NOT_SENT",
                            "attempts_used": 0,
                            "receipt_url": None,
                            "cloudinary_public_id": None,
                            "reject_reason": None,
                            "last_payment_id": None,
                            "updated_at": now,
                        },
                        "$unset": {
                            "cooldown_expiry": "",
                            "second_rejection_at": ""
                        }
                    }
                )

                # Insert log
                promotion_logs.insert_one({
                    "student_id": student_id,
                    "previous_semester": current_sem,
                    "previous_year": current_year,
                    "new_semester": next_sem,
                    "new_year": new_year,
                    "student_type": student_type,
                    "completion_time": now,
                    "promotion_time": now,
                    "no_due_cleared": no_due_cleared,
                })
                promoted_count += 1

                # 🚀 Broadcast real-time promotion to the specific student
                try:
                    notify_student(student_id, "student_promoted", {
                        "new_semester": next_sem,
                        "new_year": new_year,
                        "status": "NOT_SENT",
                        "attempts_used": 0,
                        "attempts_remaining": 2,
                    })
                except Exception as be:
                    logger.warning("[WS] Failed to broadcast promotion to student %s: %s", student_id, be)

            audit_logger.info("PROMOTION completed: from_sem=%s promoted=%d",
                              from_sem if from_sem is not None else "ALL", promoted_count)
            messages.success(request, f"Successfully promoted {promoted_count} students to the next semester!")

            # 🚀 Broadcast promotion completion to faculty and all offices
            try:
                notify_faculty("promotion_completed", {"promoted_count": promoted_count})
                notify_all_offices("cohort_promoted", {"promoted_count": promoted_count})
            except Exception as be:
                logger.warning("[WS] Failed to broadcast promotion completion: %s", be)

        except Exception as exc:
            # Lock is still released by `finally`; surface a safe error.
            logger.exception("Promotion failed")
            messages.error(request, "Promotion failed due to an internal error. No further changes were made.")
        finally:
            portal_settings.update_one(
                {"_id": "global_config"},
                {"$set": {"promotion_in_progress": False}}
            )

    return redirect("faculty_promotion")


@institution_login_required
def remove_sem8_students(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if not request.session.get("promotion_unlocked"):
        messages.error(request, "Access denied. Please verify password first.")
        return redirect("faculty_promotion")

    if request.method == "POST":
        query = {"semester": 8}
        students = list(students_col.find(query))
        if not students:
            messages.warning(request, "No Semester 8 students found to remove.")
            return redirect("faculty_promotion")

        student_ids = [s["_id"] for s in students]

        students_col.delete_many({"_id": {"$in": student_ids}})
        no_due_col.delete_many({"student_id": {"$in": student_ids}})

        from .mongo import promotion_logs
        promotion_logs.delete_many({"student_id": {"$in": student_ids}})

        messages.success(request, f"Successfully removed {len(students)} Semester 8 students from the system.")

        try:
            notify_faculty("student_roster_updated", {})
            notify_all_offices("student_roster_updated", {})
        except Exception as be:
            logger.warning("[WS] Failed to broadcast sem8 removal: %s", be)

    return redirect("faculty_promotion")



@institution_login_required
def download_student_template(request):
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    from openpyxl.styles import numbers as xl_numbers
    from openpyxl.cell.cell import TYPE_STRING

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Template"

    # ── Column widths ─────────────────────────────────────
    ws.column_dimensions["A"].width = 15   # roll_no
    ws.column_dimensions["B"].width = 20   # reg_no
    ws.column_dimensions["C"].width = 25   # name
    ws.column_dimensions["D"].width = 15   # dob
    ws.column_dimensions["E"].width = 15   # phone
    ws.column_dimensions["F"].width = 12   # semester

    # ── Apply cell formats for rows 1-500 BEFORE writing any data ────────────
    # "@" = Text  →  prevents 830123104032 becoming 8.31E+11
    #              →  prevents 9876543210  becoming a float
    for row in range(1, 501):
        ws[f"A{row}"].number_format = "@"           # roll_no  → Text
        ws[f"B{row}"].number_format = "@"           # reg_no   → Text
        ws[f"E{row}"].number_format = "@"           # phone    → Text
        ws[f"D{row}"].number_format = "yyyy-mm-dd"  # dob      → Date

    # ── Header row (row 1) ────────────────────────────────
    for col, header in enumerate(["roll_no", "reg_no", "name", "dob", "phone", "semester"], start=1):
        ws.cell(row=1, column=col).value = header

    # ── Sample row (row 2) — all text columns stored as strings ──────────────
    def write_text(row, col, value):
        """Write value as explicit Text string so Excel stores it as-is."""
        cell = ws.cell(row=row, column=col)
        cell.value = str(value)
        cell.data_type = TYPE_STRING   # force Excel to treat as Text

    write_text(2, 1, "21CS001")        # roll_no
    write_text(2, 2, "202110000001")   # reg_no  (12 digits, no scientific notation)
    write_text(2, 3, "SAMPLE NAME")    # name
    ws.cell(row=2, column=4).value = "2003-01-01"   # dob  (date string)
    write_text(2, 5, "9876543210")     # phone   (10 digits, no conversion)
    ws.cell(row=2, column=6).value = 5              # semester (number is fine)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_template.xlsx"'

    wb.save(response)
    return response


def _parse_excel_dob(dob_raw):
    """
    Parses and normalizes date of birth values from Excel into YYYY-MM-DD format.
    Handles datetime objects, date objects, Excel serial dates, and common date string formats.
    """
    if dob_raw is None:
        return ""
    if isinstance(dob_raw, (datetime, date)):
        return dob_raw.strftime("%Y-%m-%d")

    # Handle numeric serial numbers from Excel (e.g. 38450)
    if isinstance(dob_raw, (int, float)):
        try:
            converted_date = date(1899, 12, 30) + timedelta(days=int(dob_raw))
            return converted_date.strftime("%Y-%m-%d")
        except Exception:
            pass

    dob_str = str(dob_raw).strip()
    if not dob_str:
        return ""

    # Try common date formats
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(dob_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return dob_str


def _clean_excel_str(val):
    """Clean string from Excel cell, stripping trailing float decimals (e.g., 9876543210.0 -> 9876543210)."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


@institution_login_required
def import_students_excel(request):
    # 🔐 ROLE CHECK
    if request.session.get("role") != "FACULTY":
        return redirect("index")

    if request.method != "POST":
        return redirect("faculty_dashboard")

    excel = request.FILES.get("excel")
    branch = request.POST.get("branch")
    year = request.POST.get("year")

    if not excel:
        messages.error(request, "No Excel file uploaded ❌")
        return redirect(f"/faculty/?branch={branch}&year={year}")

    inserted = 0
    skipped = 0
    skipped_students = []   # name + reason

    try:
        year_int = int(year)
    except (ValueError, TypeError):
        messages.error(request, "Invalid year selected for import ❌")
        return redirect(f"/faculty/?branch={branch}&year={year}")

    try:
        wb = load_workbook(excel, read_only=True, data_only=True)
        ws = wb.active

        # ── Pass 1: parse, validate & clean every row into candidates ──
        candidates = []          # dicts ready to insert
        seen_in_file = set()     # (roll_no or reg_no) already seen this file
        roll_nos = set()
        reg_nos = set()

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 6:
                # blank tail rows in read_only mode → ignore silently
                if row and any(row):
                    skipped += 1
                    skipped_students.append(f"Row {idx} (Invalid column format)")
                continue

            roll_no, reg_no, name, dob, phone, semester = row[:6]

            if not any([roll_no, reg_no, name, dob, phone, semester]):
                continue

            roll_no = _clean_excel_str(roll_no).upper()
            reg_no = _clean_excel_str(reg_no)
            name = _clean_excel_str(name).upper()
            phone = _clean_excel_str(phone)
            dob = _parse_excel_dob(dob)

            if not roll_no or not reg_no or not name:
                skipped += 1
                skipped_students.append(
                    f"{name or 'Unknown'} (Row {idx} – Missing RegNo / RollNo)"
                )
                continue

            if not dob:
                skipped += 1
                skipped_students.append(
                    f"{name} (Row {idx} – Missing Date of Birth)"
                )
                continue

            try:
                semester_int = int(semester)
            except (ValueError, TypeError):
                skipped += 1
                skipped_students.append(f"{name} (Row {idx} – Invalid semester)")
                continue

            # Duplicate WITHIN the uploaded file.
            if roll_no in seen_in_file or reg_no in seen_in_file:
                skipped += 1
                skipped_students.append(f"{name} (Duplicate in file)")
                continue
            seen_in_file.add(roll_no)
            seen_in_file.add(reg_no)

            roll_nos.add(roll_no)
            reg_nos.add(reg_no)
            candidates.append({
                "roll_no": roll_no,
                "reg_no": reg_no,
                "name": name,
                "dob": dob,
                "phone": phone,
                "branch": branch,
                "year": year_int,
                "semester": semester_int,
                "student_type": "Hosteller",
            })

        wb.close()

        # ── One query for ALL existing duplicates (was one per row) ──
        existing_rolls = set()
        existing_regs = set()
        if roll_nos or reg_nos:
            for doc in students_col.find(
                {"$or": [
                    {"roll_no": {"$in": list(roll_nos)}},
                    {"reg_no": {"$in": list(reg_nos)}},
                ]},
                {"roll_no": 1, "reg_no": 1},
            ):
                if doc.get("roll_no"):
                    existing_rolls.add(doc["roll_no"])
                if doc.get("reg_no"):
                    existing_regs.add(doc["reg_no"])

        to_insert = []
        for c in candidates:
            if c["roll_no"] in existing_rolls or c["reg_no"] in existing_regs:
                skipped += 1
                skipped_students.append(f"{c['name']} (Duplicate RegNo / RollNo)")
            else:
                to_insert.append(c)

        # ── One bulk insert (unordered: a bad doc won't abort the rest) ──
        if to_insert:
            try:
                result = students_col.insert_many(to_insert, ordered=False)
                inserted = len(result.inserted_ids)
            except Exception as bulk_exc:
                # Unique-index collision on a concurrent import → count what stuck.
                inserted = getattr(bulk_exc, "details", {}).get("nInserted", 0) \
                    if hasattr(bulk_exc, "details") else 0
                skipped += len(to_insert) - inserted
                logger.warning("Bulk import partial failure: %s", bulk_exc)

        messages.success(
            request,
            f"Excel Import Completed ✅ Added: {inserted}, Skipped: {skipped}"
        )
        if skipped_students:
            messages.warning(
                request,
                "Skipped Students:\n" + "\n".join(skipped_students)
            )
        if inserted > 0:
            try:
                notify_faculty("student_roster_updated", {"branch": branch, "year": year})
                notify_all_offices("student_roster_updated", {"branch": branch, "year": year})
            except Exception as be:
                logger.warning("[WS] Failed to broadcast import: %s", be)

        logger.info("Excel import branch=%s year=%s added=%d skipped=%d",
                    branch, year, inserted, skipped)

    except Exception as e:
        logger.exception("Excel import failed")
        from pymongo.errors import PyMongoError, ServerSelectionTimeoutError, AutoReconnect
        if isinstance(e, (PyMongoError, ServerSelectionTimeoutError, AutoReconnect)):
            messages.error(
                request,
                "Excel import failed ❌ Database connection error. Please verify your MongoDB connection and MONGO_URI in .env."
            )
        else:
            messages.error(request, f"Excel import failed ❌ {str(e)}")

    return redirect(f"/faculty/?branch={branch}&year={year}")

# ================= LOGOUT =================
def logout_view(request):
    """
    Invalidates the authentication session completely, removes cookies,
    sets strict anti-caching headers, and redirects to login page.
    """
    request.session.flush()
    response = redirect("index")
    # Explicitly clear session & CSRF cookies from client browser
    response.delete_cookie(settings.SESSION_COOKIE_NAME, path="/")
    response.delete_cookie(settings.CSRF_COOKIE_NAME, path="/")
    response["Cache-Control"] = "no-cache, no-store, must-revalidate, private, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


# ================= OFFICE STUDENT STATUS API =================
from django.http import JsonResponse
import math

@institution_login_required
def office_student_status_api(request):
    role = request.session.get("role")
    if role not in ("LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT"):
        return JsonResponse({"error": "Unauthorized"}, status=403)

    try:
        year = int(request.GET.get("year", 0))
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid Year"}, status=400)

    if not year:
        return JsonResponse({"error": "Year is required"}, status=400)

    search = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "All").strip()
    branch_filter = request.GET.get("branch", "").strip()
    
    try:
        page = int(request.GET.get("page", 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1

    limit = 10  # number of students per page

    # Build match query for students
    match_query = {
        "year": year
    }

    if role == "DEPARTMENT":
        match_query["branch"] = request.session.get("department")
    elif branch_filter:
        match_query["branch"] = branch_filter

    if search:
        match_query["$or"] = [
            {"reg_no": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}}
        ]

    # Build aggregation pipeline
    pipeline = [
        {"$match": match_query},
        {
            "$lookup": {
                "from": "no_due_requests",
                "let": {"student_id": "$_id"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$student_id", "$$student_id"]},
                                    {"$eq": ["$office", role]}
                                ]
                            }
                        }
                    }
                ],
                "as": "no_due_record"
            }
        },
        {
            "$addFields": {
                "no_due": {"$arrayElemAt": ["$no_due_record", 0]}
            }
        },
        {
            "$addFields": {
                "office_status": {
                    "$cond": {
                        "if": {"$and": [{"$eq": [role, "HOSTEL"]}, {"$eq": ["$student_type", "Day Scholar"]}]},
                        "then": "Completed",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$no_due.status", "APPROVED"]},
                                "then": "Completed",
                                "else": {
                                    "$cond": {
                                        "if": {"$eq": ["$no_due.status", "PENDING"]},
                                        "then": "Pending",
                                        "else": "Incomplete"
                                    }
                                }
                            }
                        }
                    }
                },
                "completed_time": {
                    "$cond": {
                        "if": {"$and": [{"$eq": [role, "HOSTEL"]}, {"$eq": ["$student_type", "Day Scholar"]}]},
                        "then": "-",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$no_due.status", "APPROVED"]},
                                "then": {"$ifNull": ["$no_due.updated_at", "$no_due.created_at"]},
                                "else": "-"
                            }
                        }
                    }
                }
            }
        }
    ]

    # Apply status filter
    if status_filter and status_filter != "All":
        pipeline.append({"$match": {"office_status": status_filter}})

    # Pagination facet
    pipeline.append({
        "$facet": {
            "metadata": [{"$count": "total"}],
            "data": [
                {"$skip": (page - 1) * limit},
                {"$limit": limit}
            ]
        }
    })

    try:
        results = list(students_col.aggregate(pipeline))
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    total_count = 0
    students_list = []

    if results:
        metadata = results[0].get("metadata", [])
        if metadata:
            total_count = metadata[0]["total"]
        data = results[0].get("data", [])
        
        for s in data:
            ct = s.get("completed_time")
            # Format time if it is a datetime object
            if isinstance(ct, datetime):
                completed_time_str = ct.strftime("%d-%m-%Y %I:%M %p")
            else:
                completed_time_str = "-"
                
            students_list.append({
                "student_id": str(s.get("_id", "")),
                "reg_no": s.get("reg_no"),
                "name": s.get("name"),
                "branch": s.get("branch"),
                "semester": s.get("semester"),
                "is_75_scheme": s.get("is_75_scheme", False),
                "status": s.get("office_status"),
                "completed_time": completed_time_str
            })

    total_pages = math.ceil(total_count / limit)

    return JsonResponse({
        "students": students_list,
        "current_page": page,
        "total_pages": total_pages,
        "total_count": total_count
    })


# ═══════════════════════════════════════════
#  REPORT GENERATION & EXPORT VIEWS
# ═══════════════════════════════════════════

def _get_report_students_pipeline(role, report_type, year_str, branch_val, dept):
    match_query = {}

    if role == "DEPARTMENT":
        match_query["branch"] = dept
        branch_val = dept

    if report_type == "year":
        if not year_str:
            return None, "Year is required for Year Wise Report"
        try:
            match_query["year"] = int(year_str)
        except (ValueError, TypeError):
            return None, "Invalid Year"

    elif report_type == "branch":
        if role != "DEPARTMENT":
            if not branch_val:
                return None, "Branch is required for Branch Wise Report"
            match_query["branch"] = branch_val

    elif report_type == "year_branch":
        if not year_str:
            return None, "Year is required for Year + Branch Wise Report"
        try:
            match_query["year"] = int(year_str)
        except (ValueError, TypeError):
            return None, "Invalid Year"
        if role != "DEPARTMENT":
            if not branch_val:
                return None, "Branch is required for Year + Branch Wise Report"
            match_query["branch"] = branch_val

    pipeline = [
        {"$match": match_query},
        {
            "$lookup": {
                "from": "no_due_requests",
                "let": {"student_id": "$_id"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$student_id", "$$student_id"]},
                                    {"$eq": ["$office", role]}
                                ]
                            }
                        }
                    }
                ],
                "as": "no_due_record"
            }
        },
        {
            "$addFields": {
                "no_due": {"$arrayElemAt": ["$no_due_record", 0]}
            }
        },
        {
            "$addFields": {
                "office_status": {
                    "$cond": {
                        "if": {"$and": [{"$eq": [role, "HOSTEL"]}, {"$eq": ["$student_type", "Day Scholar"]}]},
                        "then": "Completed",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$no_due.status", "APPROVED"]},
                                "then": "Completed",
                                "else": {
                                    "$cond": {
                                        "if": {"$eq": ["$no_due.status", "PENDING"]},
                                        "then": "Pending",
                                        "else": "Incomplete"
                                    }
                                }
                            }
                        }
                    }
                },
                "completed_time": {
                    "$cond": {
                        "if": {"$and": [{"$eq": [role, "HOSTEL"]}, {"$eq": ["$student_type", "Day Scholar"]}]},
                        "then": "-",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$no_due.status", "APPROVED"]},
                                "then": {"$ifNull": ["$no_due.updated_at", "$no_due.created_at"]},
                                "else": "-"
                            }
                        }
                    }
                }
            }
        },
        {"$sort": {"reg_no": 1}}
    ]
    return pipeline, None


@institution_login_required
def office_report_preview_api(request):
    role = request.session.get("role")
    if role not in ("LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT"):
        return JsonResponse({"error": "Unauthorized"}, status=403)

    report_type = request.GET.get("report_type", "").strip()
    if report_type not in ("year", "branch", "year_branch"):
        return JsonResponse({"error": "Invalid Report Type"}, status=400)

    year_str = request.GET.get("year", "").strip()
    branch_val = request.GET.get("branch", "").strip()
    dept = request.session.get("department")

    pipeline, err = _get_report_students_pipeline(role, report_type, year_str, branch_val, dept)
    if err:
        return JsonResponse({"error": err}, status=400)

    try:
        page = int(request.GET.get("page", 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1

    limit = 10
    pagination_pipeline = list(pipeline)
    pagination_pipeline.append({
        "$facet": {
            "metadata": [{"$count": "total"}],
            "data": [
                {"$skip": (page - 1) * limit},
                {"$limit": limit}
            ]
        }
    })

    try:
        results = list(students_col.aggregate(pagination_pipeline))
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    total_count = 0
    students_list = []

    if results:
        metadata = results[0].get("metadata", [])
        if metadata:
            total_count = metadata[0]["total"]
        data = results[0].get("data", [])

        for s in data:
            ct = s.get("completed_time")
            if isinstance(ct, datetime):
                completed_time_str = ct.strftime("%d-%m-%Y %I:%M %p")
            else:
                completed_time_str = "-"

            students_list.append({
                "reg_no": s.get("reg_no"),
                "roll_no": s.get("roll_no", ""),
                "name": s.get("name"),
                "branch": s.get("branch"),
                "year": s.get("year"),
                "semester": s.get("semester"),
                "student_type": s.get("student_type", "Hosteller"),
                "is_75_scheme": s.get("is_75_scheme", False),
                "status": s.get("office_status"),
                "completed_time": completed_time_str
            })

    total_pages = math.ceil(total_count / limit)

    return JsonResponse({
        "students": students_list,
        "current_page": page,
        "total_pages": total_pages,
        "total_count": total_count
    })


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#4B5563"))
        
        # Line above footer
        width, height = A4
        self.setStrokeColor(colors.HexColor("#D1D5DB"))
        self.setLineWidth(0.5)
        self.line(36, 45, width - 36, 45)
        
        # Footer content
        self.drawString(36, 32, "Generated by GCES No Due Clearance Portal")
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(width - 36, 32, page_text)
        self.restoreState()


@institution_login_required
def office_report_pdf_view(request):
    role = request.session.get("role")
    if role not in ("LIBRARY", "HOSTEL", "COLLEGE", "DEPARTMENT"):
        return HttpResponse("Unauthorized", status=403)

    report_type = request.GET.get("report_type", "").strip()
    if report_type not in ("year", "branch", "year_branch"):
        return HttpResponse("Invalid Report Type", status=400)

    year_str = request.GET.get("year", "").strip()
    branch_val = request.GET.get("branch", "").strip()
    dept = request.session.get("department")

    pipeline, err = _get_report_students_pipeline(role, report_type, year_str, branch_val, dept)
    if err:
        return HttpResponse(err, status=400)

    try:
        students = list(students_col.aggregate(pipeline))
    except Exception as e:
        return HttpResponse(f"Database error: {str(e)}", status=500)

    # Prepare response as PDF attachment
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Clearance_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Set download token cookie if provided
    download_token = request.GET.get("download_token")
    if download_token:
        response.set_cookie("fileDownloadToken", download_token, max_age=60)

    # Document template setup (A4 standard: 595.27 x 841.89 points)
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=55
    )

    story = []
    styles = getSampleStyleSheet()

    primary_color = colors.HexColor("#d52b1e")
    dark_gray = colors.HexColor("#1F2937")

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=primary_color,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=dark_gray,
        alignment=1
    )

    info_label_style = ParagraphStyle(
        'InfoLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#374151")
    )

    info_val_style = ParagraphStyle(
        'InfoVal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#4B5563")
    )

    # Title Block
    story.append(Paragraph("GOVERNMENT COLLEGE OF ENGINEERING – SRIRANGAM", title_style))
    story.append(Spacer(1, 4))

    office_display = role.title() + " Office" if role != "DEPARTMENT" else f"{dept} Department Office"
    story.append(Paragraph(f"{office_display} – Clearance Status Report", subtitle_style))
    story.append(Spacer(1, 10))

    # Red divider line
    divider = Table([[""]], colWidths=[523], rowHeights=[2])
    divider.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), primary_color),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(divider)
    story.append(Spacer(1, 12))

    # Meta-info block
    now_str = datetime.now().strftime("%d-%m-%Y %I:%M %p")

    selected_year = f"Year {year_str}" if year_str else "All Years"
    selected_branch = branch_val if branch_val else "All Branches"

    report_title_display = "Clearance Report"
    if report_type == "year":
        report_title_display = "Year-Wise Clearance Report"
    elif report_type == "branch":
        report_title_display = "Branch-Wise Clearance Report"
    elif report_type == "year_branch":
        report_title_display = "Year & Branch Clearance Report"

    info_data = [
        [
            Paragraph("Report Type:", info_label_style), Paragraph(report_title_display, info_val_style),
            Paragraph("Generated on:", info_label_style), Paragraph(now_str, info_val_style)
        ],
        [
            Paragraph("Selected Year:", info_label_style), Paragraph(selected_year, info_val_style),
            Paragraph("Selected Branch:", info_label_style), Paragraph(selected_branch, info_val_style)
        ],
        [
            Paragraph("Total Students:", info_label_style), Paragraph(str(len(students)), info_val_style),
            "", ""
        ]
    ]

    info_table = Table(info_data, colWidths=[100, 160, 100, 163])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 15))

    # Student Data Table
    th_style = ParagraphStyle(
        'TableHead',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )

    td_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=dark_gray
    )

    td_completed_style = ParagraphStyle('TableBodyCompleted', parent=td_style, textColor=colors.HexColor("#16A34A"))
    td_pending_style = ParagraphStyle('TableBodyPending', parent=td_style, textColor=colors.HexColor("#D97706"))
    td_incomplete_style = ParagraphStyle('TableBodyIncomplete', parent=td_style, textColor=colors.HexColor("#DC2626"))

    table_data = [[
        Paragraph("Register No", th_style),
        Paragraph("Roll No", th_style),
        Paragraph("Student Name", th_style),
        Paragraph("Branch", th_style),
        Paragraph("Year", th_style),
        Paragraph("Sem", th_style),
        Paragraph("Type", th_style),
        Paragraph("Status", th_style),
        Paragraph("Approval Date & Time", th_style)
    ]]

    for s in students:
        ct = s.get("completed_time")
        if isinstance(ct, datetime):
            completed_time_str = ct.strftime("%d-%m-%Y %I:%M %p")
        else:
            completed_time_str = "-"

        status = s.get("office_status")
        if status == "Completed":
            status_p = Paragraph("Completed", td_completed_style)
        elif status == "Pending":
            status_p = Paragraph("Pending", td_pending_style)
        else:
            status_p = Paragraph("Incomplete", td_incomplete_style)

        table_data.append([
            Paragraph(str(s.get("reg_no", "")), td_style),
            Paragraph(str(s.get("roll_no", "")), td_style),
            Paragraph(str(s.get("name", "")), td_style),
            Paragraph(str(s.get("branch", "")), td_style),
            Paragraph(str(s.get("year", "")), td_style),
            Paragraph(str(s.get("semester", "")), td_style),
            Paragraph(str(s.get("student_type", "Hosteller")), td_style),
            status_p,
            Paragraph(completed_time_str, td_style)
        ])

    student_table = Table(table_data, colWidths=[75, 55, 95, 50, 25, 25, 60, 60, 78], repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
    ])

    for i in range(1, len(table_data)):
        bg_color = colors.HexColor("#F9FAFB") if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg_color)

    student_table.setStyle(t_style)
    story.append(student_table)

    doc.build(story, canvasmaker=NumberedCanvas)
    return response

